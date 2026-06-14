"""Single source of truth for the SSA portfolio, read FRESH from the workbook
every run. Trust nothing from a prior parse.

Reads:
  - Col A  Share Name      (+ FONT COLOUR: red FFFF0000 => MTF/leveraged)
  - Col B  % of Portfolio
  - Col C  Avg. Cost (INR)
  - Col E/F/G  Stocks / Quantity / Valuation in INR  (EXACT, name-keyed; this
    block is sorted by value and offset from the A/B/C block, so we join on NAME)

Only the ticker symbol, theme tag and defence-cluster membership are curated
metadata (a name->ticker map); every number comes from the file.
"""
from __future__ import annotations
import re
from openpyxl import load_workbook

XLSX = r"D:\SSA Portfolio.xlsx"

# curated metadata only: file-name -> (NSE ticker, theme, in_defence_cluster)
META = {
 "Bharat Dynamics Ltd": ("BDL.NS","Defence",True),
 "Bharat Electronics Ltd": ("BEL.NS","Defence",True),
 "Hindustan Aeronautics Ltd": ("HAL.NS","Defence",True),
 "HBL Engineering Ltd": ("HBLENGINE.NS","Defence",True),
 "Tata Power Company Ltd": ("TATAPOWER.NS","Power/Utilities",False),
 "Avantel Ltd": ("AVANTEL.NS","Defence",True),
 "Power Finance Corporation Ltd": ("PFC.NS","Financials/PSU",False),
 "Transrail Lighting Ltd": ("TRANSRAILL.NS","Capex/Infra",False),
 "Jupiter Wagons Ltd": ("JWL.NS","Capex/Infra",False),
 "Easy Trip Planners Ltd": ("EASEMYTRIP.NS","Travel/Consumer",False),
 "ACC Ltd": ("ACC.NS","Cement/Materials",False),
 "Indian Renewable Energy Development Agency Ltd": ("IREDA.NS","Financials/PSU",False),
 "Apollo Micro Systems Ltd": ("APOLLO.NS","Defence",True),
 "Balaji Amines Ltd": ("BALAMINES.NS","Chemicals",False),
 "Piramal Pharma Ltd": ("PPLPHARMA.NS","Pharma",False),
 "Maharashtra Seamless Ltd": ("MAHSEAMLES.NS","Capex/Infra",False),
 "Olectra Greentech Ltd": ("OLECTRA.NS","EV/Capital Goods",False),
 "RattanIndia Power Ltd": ("RTNPOWER.NS","Power/Distressed",False),
 "Astra Microwave Products Ltd": ("ASTRAMICRO.NS","Defence",True),
 "Happiest Minds Technologies Ltd": ("HAPPSTMNDS.NS","IT",False),
 "Gujarat Pipavav Port Ltd": ("GPPL.NS","Capex/Infra",False),
 "Tata Technologies Ltd": ("TATATECH.NS","IT",False),
 "Avalon Technologies Ltd": ("AVALON.NS","Defence/EMS",True),
 "GMM Pfaudler Ltd": ("GMMPFAUDLR.NS","Chemicals/CapGoods",False),
 "GTL Infrastructure Ltd": ("GTLINFRA.NS","Telecom/Distressed",False),
 "Akshar Spintex Ltd": ("AKSHAR.NS","Textiles/Distressed",False),
 "HFCL Ltd": ("HFCL.NS","Telecom",False),
 "Ajooni Biotech Ltd": ("AJOONI.NS","Distressed",False),
 "Ircon International Ltd": ("IRCON.NS","Capex/Infra",False),
 "Future Consumer Ltd": ("FCONSUMER.NS","Distressed",False),
 "TD Power Systems Ltd": ("TDPOWERSYS.NS","Capex/Infra",False),
 "Reliance Communications Ltd": ("RCOM.NS","Telecom/Distressed",False),
 "Kshitij Polyline Ltd": ("KSHITIJPOL.NS","Distressed",False),
 "Housing Development & Infrastructure Ltd": ("HDIL.NS","RealEstate/Distressed",False),
 "Vikas Lifecare Ltd": ("VIKASLIFE.NS","Distressed",False),
 "Vikas Ecotech Ltd": ("VIKASECO.NS","Distressed",False),
 "Future Enterprises Ltd": ("FEL.NS","Distressed",False),
 "Rail Vikas Nigam Ltd": ("RVNL.NS","Capex/Infra",False),
}

RED = {"FFFF0000", "FF0000"}


def load_holdings() -> list[dict]:
    """Return list of holding dicts, freshly parsed. Raises if structure drifts."""
    wb = load_workbook(XLSX)                  # formatting (font colour)
    wbd = load_workbook(XLSX, data_only=True) # cached values
    ws, wsd = wb.active, wbd.active

    # A/B/C block: name -> (weight, cost, is_mtf)
    abc = {}
    mtf_names = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        name = str(name).strip()
        weight = wsd.cell(r, 2).value
        cost = wsd.cell(r, 3).value
        rgb = getattr(getattr(ws.cell(r, 1).font, "color", None), "rgb", None)
        is_mtf = isinstance(rgb, str) and rgb.upper() in RED
        if is_mtf:
            mtf_names.append(name)
        abc[name] = {
            "weight": float(weight) if isinstance(weight, (int, float)) else None,
            "avg_cost": float(cost) if isinstance(cost, (int, float)) else None,
            "mtf": is_mtf,
        }

    # E/F/G block: name -> (qty, file_valuation)  [name-keyed; offset block]
    efg = {}
    for r in range(2, wsd.max_row + 1):
        nm = wsd.cell(r, 5).value
        if not nm:
            continue
        nm = str(nm).strip()
        qty = wsd.cell(r, 6).value
        val = wsd.cell(r, 7).value
        efg[nm] = {
            "qty": float(qty) if isinstance(qty, (int, float)) else None,
            "file_valuation": float(val) if isinstance(val, (int, float)) else None,
        }

    holdings = []
    for name, m in abc.items():
        if name not in META:
            raise SystemExit(f"Unmapped holding (add to META): {name!r}")
        tk, theme, defence = META[name]
        e = efg.get(name, {})
        holdings.append({
            "ticker": tk, "name": name, "theme": theme,
            "in_defence_cluster": defence,
            "weight": m["weight"], "avg_cost": m["avg_cost"], "mtf": m["mtf"],
            "qty": e.get("qty"), "file_valuation": e.get("file_valuation"),
        })

    n_mtf = sum(1 for h in holdings if h["mtf"])
    if n_mtf != 8:
        raise SystemExit(f"MTF font-colour read returned {n_mtf} (expected 8): {mtf_names}")
    if len(holdings) != 38:
        raise SystemExit(f"Expected 38 holdings, got {len(holdings)}")
    return holdings


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")
    hs = load_holdings()
    tot = sum(h["file_valuation"] or 0 for h in hs)
    mtf = sum(h["file_valuation"] or 0 for h in hs if h["mtf"])
    print(f"{len(hs)} holdings | file total ₹{tot:,.0f} | MTF ₹{mtf:,.0f} | cash ₹{tot-mtf:,.0f}")
    print(f"MTF ({sum(1 for h in hs if h['mtf'])}):")
    for h in hs:
        if h["mtf"]:
            print(f"   {h['name']:<30} qty={h['qty']:>8.0f}  fileVal ₹{h['file_valuation']:>12,.0f}  cost ₹{h['avg_cost']}")
