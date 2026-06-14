"""Self-funding MTF de-leverage plan, computed from EXACT file values.
Reads D:\\SSA Portfolio.xlsx columns E (Stocks) / F (Quantity) / G (Valuation in
INR) -- keyed by NAME in column E (that block is sorted differently from A/B/C).
Target = current market value of the 8 MTF holdings; funded by selling profitable
CASH holdings only, in a disciplined priority order. No new capital."""
from __future__ import annotations
import json, math, sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

NAME2TK = {
 "Bharat Dynamics Ltd":"BDL.NS","Bharat Electronics Ltd":"BEL.NS","Hindustan Aeronautics Ltd":"HAL.NS",
 "HBL Engineering Ltd":"HBLENGINE.NS","Tata Power Company Ltd":"TATAPOWER.NS","Avantel Ltd":"AVANTEL.NS",
 "Power Finance Corporation Ltd":"PFC.NS","Transrail Lighting Ltd":"TRANSRAILL.NS","Jupiter Wagons Ltd":"JWL.NS",
 "Easy Trip Planners Ltd":"EASEMYTRIP.NS","ACC Ltd":"ACC.NS",
 "Indian Renewable Energy Development Agency Ltd":"IREDA.NS","Apollo Micro Systems Ltd":"APOLLO.NS",
 "Balaji Amines Ltd":"BALAMINES.NS","Piramal Pharma Ltd":"PPLPHARMA.NS","Maharashtra Seamless Ltd":"MAHSEAMLES.NS",
 "Olectra Greentech Ltd":"OLECTRA.NS","RattanIndia Power Ltd":"RTNPOWER.NS","Astra Microwave Products Ltd":"ASTRAMICRO.NS",
 "Happiest Minds Technologies Ltd":"HAPPSTMNDS.NS","Gujarat Pipavav Port Ltd":"GPPL.NS","Tata Technologies Ltd":"TATATECH.NS",
 "Avalon Technologies Ltd":"AVALON.NS","GMM Pfaudler Ltd":"GMMPFAUDLR.NS","GTL Infrastructure Ltd":"GTLINFRA.NS",
 "Akshar Spintex Ltd":"AKSHAR.NS","HFCL Ltd":"HFCL.NS","Ajooni Biotech Ltd":"AJOONI.NS","Ircon International Ltd":"IRCON.NS",
 "Future Consumer Ltd":"FCONSUMER.NS","TD Power Systems Ltd":"TDPOWERSYS.NS","Reliance Communications Ltd":"RCOM.NS",
 "Kshitij Polyline Ltd":"KSHITIJPOL.NS","Housing Development & Infrastructure Ltd":"HDIL.NS","Vikas Lifecare Ltd":"VIKASLIFE.NS",
 "Vikas Ecotech Ltd":"VIKASECO.NS","Future Enterprises Ltd":"FEL.NS","Rail Vikas Nigam Ltd":"RVNL.NS",
}
MTF = {"PFC.NS","TRANSRAILL.NS","JWL.NS","EASEMYTRIP.NS","ACC.NS","BALAMINES.NS","PPLPHARMA.NS","MAHSEAMLES.NS"}
# avg cost lives in column C keyed by column-A name; build that map too.
ws = load_workbook(r"D:\SSA Portfolio.xlsx", data_only=True).active
avgcost = {}
for r in range(2, ws.max_row+1):
    a, c = ws.cell(r,1).value, ws.cell(r,3).value
    if a and c is not None:
        avgcost[str(a).strip()] = float(c)

# read E (name) / F (qty) / G (value) keyed by name
pos = {}
for r in range(2, ws.max_row+1):
    e, f, g = ws.cell(r,5).value, ws.cell(r,6).value, ws.cell(r,7).value
    if not e:
        continue
    nm = str(e).strip()
    tk = NAME2TK.get(nm)
    qty = float(f) if f is not None else None
    val = float(g) if g is not None else None
    cost = avgcost.get(nm)
    pos[tk] = {"name": nm, "ticker": tk, "qty": qty, "value": val,
               "avg_cost": cost, "mtf": tk in MTF,
               "cost_basis": (qty*cost if qty and cost is not None else None)}
    if val and pos[tk]["cost_basis"] is not None:
        pos[tk]["gain"] = val - pos[tk]["cost_basis"]
        pos[tk]["gain_pct"] = round(pos[tk]["gain"]/pos[tk]["cost_basis"]*100,1) if pos[tk]["cost_basis"] else None
        pos[tk]["price"] = val/qty

total = sum(p["value"] for p in pos.values() if p["value"])
mtf_val = sum(p["value"] for p in pos.values() if p["mtf"] and p["value"])
cash_val = total - mtf_val
print(f"Total portfolio value (sum col G): ₹{total:,.0f}")
print(f"MTF book value (8 holdings):       ₹{mtf_val:,.0f}  <- TARGET to raise")
print(f"Cash book value:                   ₹{cash_val:,.0f}")
print("\nMTF holdings (the target):")
for tk in sorted(MTF, key=lambda t:-pos[t]["value"]):
    p = pos[tk]; print(f"  {p['name']:<26} qty {int(p['qty']):>7}  val ₹{p['value']:>12,.0f}  P&L {p['gain_pct']:>6}%")

# ── disciplined sell priority (CASH holdings only) ───────────────────────────
# 1) frothy / TRIM-EXIT-flagged winners (good hygiene anyway); 2) trim the big
# doubled defence winners (raises cash AND cuts the 60% concentration).
PRIORITY = ["AVANTEL.NS","APOLLO.NS","ASTRAMICRO.NS","AVALON.NS","HFCL.NS",
            "HBLENGINE.NS","BEL.NS","HAL.NS","BDL.NS"]
REASON = {"AVANTEL.NS":"frothy + insider selling (TRIM/EXIT)","APOLLO.NS":"+frothy small-cap defence (TRIM)",
 "ASTRAMICRO.NS":"frothy small-cap defence (TRIM)","AVALON.NS":"frothy small-cap defence/EMS (TRIM)",
 "HFCL.NS":"richly-valued trade (TRIM)","HBLENGINE.NS":"doubled defence winner -> cuts concentration",
 "BEL.NS":"doubled defence winner -> cuts concentration","HAL.NS":"quality defence -> last-resort trim",
 "BDL.NS":"over-valued top weight -> last-resort trim"}

target = mtf_val
sells, raised = [], 0.0
for tk in PRIORITY:
    if raised >= target - 1: break
    p = pos[tk]
    if p["mtf"]:  # never fund by selling MTF itself
        continue
    remaining = target - raised
    price = p["price"]
    if p["value"] <= remaining + 1:   # full sell
        shares = int(p["qty"]); proceeds = p["value"]; kind = "FULL"
    else:                              # partial sell to top up
        shares = math.ceil(remaining/price); proceeds = shares*price; kind = "PARTIAL"
    cost = shares*p["avg_cost"]; gain = proceeds - cost
    raised += proceeds
    sells.append({"name":p["name"],"ticker":tk,"kind":kind,"shares_sold":shares,
                  "shares_total":int(p["qty"]),"pct_of_position":round(shares/p["qty"]*100,1),
                  "proceeds":round(proceeds),"realized_gain":round(gain),
                  "gain_pct":round(gain/cost*100,1) if cost else None,
                  "reason":REASON[tk],"cumulative":round(raised)})

net_profit = sum(s["realized_gain"] for s in sells)
surplus = raised - target
print("\nSELL LIST (cash holdings, priority order):")
for s in sells:
    print(f"  {s['name']:<22} {s['kind']:<7} {s['shares_sold']:>6}/{s['shares_total']:<6} "
          f"({s['pct_of_position']:>5}%)  proceeds ₹{s['proceeds']:>11,}  gain ₹{s['realized_gain']:>11,}  cum ₹{s['cumulative']:>11,}")
print(f"\nNET-NET:  target ₹{target:,.0f} | raised ₹{raised:,.0f} | surplus ₹{surplus:,.0f} | net profit realised ₹{net_profit:,.0f}")

json.dump({"total_portfolio_value":round(total),"mtf_book_value":round(mtf_val),
           "cash_book_value":round(cash_val),"target":round(target),"raised":round(raised),
           "surplus":round(surplus),"net_profit_realized":round(net_profit),
           "mtf_holdings":[{"name":pos[t]["name"],"value":round(pos[t]["value"]),"gain_pct":pos[t]["gain_pct"],
                            "qty":int(pos[t]["qty"])} for t in sorted(MTF,key=lambda t:-pos[t]["value"])],
           "sells":sells, "positions":pos},
          open(r"D:\liquidation_plan.json","w",encoding="utf-8"), indent=2, default=str)
print("\nSaved D:\\liquidation_plan.json")
