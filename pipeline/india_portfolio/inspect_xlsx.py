"""Fresh, trust-nothing dump of the SSA portfolio workbook.
Prints every non-empty cell with its font colour so MTF (red) vs cash (black)
is read from FORMATTING, and confirms whether exact qty / market-value columns
exist. Nothing here is reused from the earlier parse."""
import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

PATH = r"D:\SSA Portfolio.xlsx"
wb = load_workbook(PATH)            # keep formatting (font colours)
wb_d = load_workbook(PATH, data_only=True)  # cached values
print("Sheets:", wb.sheetnames)
ws = wb.active
wsd = wb_d.active
print(f"Active='{ws.title}'  dims={ws.dimensions}  max_row={ws.max_row} max_col={ws.max_col if hasattr(ws,'max_col') else ws.max_column}")

# Header row
print("\n--- Row 1 (headers) ---")
for col in range(1, ws.max_column + 1):
    v = ws.cell(1, col).value
    if v is not None:
        print(f"  col {col} ({chr(64+col)}): {v!r}")

def colour(cell):
    col = getattr(cell.font, "color", None)
    rgb = getattr(col, "rgb", None)
    theme = getattr(col, "theme", None)
    return rgb, theme

print("\n--- Column A with font colour (MTF detection) ---")
red, black = [], []
for r in range(1, ws.max_row + 1):
    c = ws.cell(r, 1)
    if c.value is None:
        continue
    rgb, theme = colour(c)
    tag = ""
    if isinstance(rgb, str) and rgb.upper() in ("FFFF0000", "FF0000"):
        tag = "  <== RED (MTF)"
        red.append(str(c.value).strip())
    else:
        black.append(str(c.value).strip())
    print(f"  A{r}: {str(c.value)[:42]:<42} rgb={rgb} theme={theme}{tag}")

print(f"\nRED (MTF) count={len(red)}:")
for x in red: print("   ", x)

print("\n--- Full grid (cols A..H), data_only values ---")
for r in range(1, ws.max_row + 1):
    row = []
    for col in range(1, min(ws.max_column, 8) + 1):
        v = wsd.cell(r, col).value
        if v is not None:
            row.append(f"{chr(64+col)}={v}")
    if row:
        print(f"  r{r}: " + " | ".join(row))
